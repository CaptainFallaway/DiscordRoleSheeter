import openpyxl
from datetime import datetime
from typing import TypeAliasType
from pydantic import BaseModel, model_validator, Field

SnowFlake = TypeAliasType("SnowFlake", str)


class Role(BaseModel):
    id: SnowFlake
    name: str
    color:  str

    @model_validator(mode="before")
    def int_to_hex(cls, values) -> dict:
        values['color'] = f"#{values['color']:06x}"
        return values


class User(BaseModel):
    id: SnowFlake
    username: str
    bot: bool = Field(default=False, alias="bot")


class Member(BaseModel):
    roles: list[str]  # NOTE maybe a list of Role objects?
    user: User


class ExcelMember(BaseModel):
    """
    Describes a member from the excel sheet

    properties:
        user_id: SnowFlake
        roles: list[str]
    """

    user_id: SnowFlake
    roles: list[str]


class MemberChanges(BaseModel):
    """
    Descibes a members changes from the excel sheet

    properties:
        user_id: SnowFlake
        added_roles: list[SnowFlake] = []
        removed_roles: list[SnowFlake] = []
    """

    user_id: SnowFlake
    added_roles: list[SnowFlake] = []
    removed_roles: list[SnowFlake] = []


workbook = openpyxl.load_workbook("checkbox_example.xlsx")

metadata_sheet = workbook["Metadata"]

metadata_date = datetime.fromisoformat(metadata_sheet["A1"].value)

metadata_role_ids: list[SnowFlake] = [
    str(_id[1].value) for _id in metadata_sheet.iter_cols() if _id[1].value is not None
    ]

metadata_members = {}
for column in metadata_sheet.iter_cols():
    user_id = str(column[2].value)
    username = column[3].value
    roles = column[4].value.split("|")

    metadata_members[username] = ExcelMember(user_id=user_id, roles=roles)

roles_sheet = workbook["Roles"]

changes = []

for row in roles_sheet.iter_rows(min_row=2):
    excel_member = metadata_members.get(row[0].value)  # get the model from the username

    if excel_member is None:
        # TODO add error that a name has changed
        continue

    member_changes = MemberChanges(user_id=excel_member.user_id)

    for i, cell in enumerate(row[1:]):
        role_id = metadata_role_ids[i]
        val = cell.value

        if val and role_id not in excel_member.roles:
            member_changes.added_roles.append(role_id)

        if not val and role_id in excel_member.roles:
            member_changes.removed_roles.append(role_id)

    if member_changes.added_roles or member_changes.removed_roles:
        changes.append(member_changes)


for change in changes:
    print(change)

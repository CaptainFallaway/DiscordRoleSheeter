import openpyxl
from datetime import datetime

from helpers.dataclasses import MemberChanges, ExcelMember, ExcelReadResponse, ErrorInfo


class ExcelReader:
    def __init__(self, excel_filename: str) -> None:
        self.excel_filename = excel_filename

    def read(self) -> ExcelReadResponse | ErrorInfo:
        try:
            workbook = openpyxl.load_workbook(self.excel_filename)
        except FileNotFoundError:
            return ErrorInfo(message="The excel file does not exist or is not readable.")

        # Get the metadata sheet data

        metadata_sheet = workbook["Metadata"]

        metadata_date = datetime.fromisoformat(metadata_sheet["A1"].value)

        metadata_role_ids = [_id[1].value for _id in metadata_sheet.iter_cols() if _id[1].value is not None]

        metadata_members = {}
        for column in metadata_sheet.iter_cols():
            if len(column) != 5:
                return ErrorInfo(message="Fatal excel sheet error. [b]Please pull again.[/b]")

            user_id = str(column[2].value)
            username = column[3].value
            roles = column[4].value.split("|")

            metadata_members[username] = ExcelMember(user_id=user_id, roles=roles)

        # Get the roles sheet data and compare it with the metadata

        roles_sheet = workbook["Roles"]

        changes = []

        for indx, row in enumerate(roles_sheet.iter_rows(min_row=2)):
            excel_member = metadata_members.get(row[0].value)  # get the model from the username

            if excel_member is None:
                return ErrorInfo(message=f"Accidental name change in cell '[b]A{indx+2}[/b]' !")

            member_changes = MemberChanges(username=row[0].value, user_id=excel_member.user_id)

            for i, cell in enumerate(row[1:]):
                role_id = metadata_role_ids[i]
                val = cell.value

                if val is None or not isinstance(val, int):
                    return ErrorInfo(message=f"Invalid cell value at '[b]{cell.column_letter}{cell.row}[/b]' !")

                if val and role_id not in excel_member.roles:
                    member_changes.added_roles.append(role_id)

                if not val and role_id in excel_member.roles:
                    member_changes.removed_roles.append(role_id)

            if member_changes.added_roles or member_changes.removed_roles:
                changes.append(member_changes)

        return ExcelReadResponse(date=metadata_date, changes=changes)
